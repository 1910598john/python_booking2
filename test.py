from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment
wb = load_workbook('customer_database.xlsx')
ws = wb.active

maxRow = ws.max_row + 1
maxCol = ws.max_column + 1
customersData_list = []
for r in range(2, maxRow):
    customer = {}
    customer['name'] = ws.cell(row=r, column=1).value
    customer['address'] = ws.cell(row=r, column=2).value
    customer['contact'] = ws.cell(row=r, column=3).value
    customer['email'] = ws.cell(row=r, column=4).value
    customer['expected_date'] = ws.cell(row=r, column=5).value
    customer['duration'] = ws.cell(row=r, column=6).value
    customer['amount_paid'] = ws.cell(row=r, column=7).value
    print(customer)
    customersData_list.append(customer)

print("      ")
def sortbydate(e):
    return e['expected_date']
customersData_list.sort(key=sortbydate)

print(customersData_list)
#wb.save('customer_database.xlsx')