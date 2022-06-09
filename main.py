#delete nyo nalang mga unnecessary comments
from tkinter import * #tkinter GUI
from openpyxl import Workbook, load_workbook #for excel automation
from openpyxl.styles.alignment import Alignment
from time import strftime
from tkinter import font

#boolean variables for windows
BOOK_WINDOW_CREATED = False #window not created
CUSTOMERS_WINDOW_CREATED = False #window not created
POP_UP_WINDOW_CREATED = False #window not created

#workbook | workbook variable
wb = load_workbook('customer_database.xlsx')
#worksheet variable
ws = wb.active
#pop-up window
def create_pop_up_window(title, text, textcolor): #function(parameters)
    global POP_UP_WINDOW_CREATED #make this variable a global variable
    if POP_UP_WINDOW_CREATED is not True: #if variable is false
        POP_UP_WINDOW_CREATED = True #set variable to true
        #create pop-up window
        pop_up_window = Toplevel() 
        #pop-up window closing function
        def close_pop_up_window():
            global POP_UP_WINDOW_CREATED #make this variable a global variable
            POP_UP_WINDOW_CREATED = False #set to false so the window can be created again
            pop_up_window.destroy() #close window
        pop_up_window.protocol('WM_DELETE_WINDOW', close_pop_up_window) #amo lang ina sya aram ko na protocol 'WM_DELETE_WINDOW' meaning pag gin close mo sya ica-call nya an 'close_pop_up_window' function
        pop_up_window.title(title) #title san window
        pop_up_window.resizable(False, False) #disable resizing
        width = 285 #window's width
        height = 70 #window's height
        #screen dimension
        screen_width = pop_up_window.winfo_screenwidth() #screen max width pixels
        screen_height = pop_up_window.winfo_screenheight() #screen max height pixels
        #kailangan ini sya para ma center an aton windows 
        center_x = int(screen_width/2 - width/2)
        center_y = int(screen_height/2 - height/2)
        #tas i-seset naton sya
        pop_up_window.geometry(f'{width}x{height}+{center_x}+{center_y}') #set book window's height and width
        pop_up_window.configure(bg='#FFFFFF')
        #message
        Label(pop_up_window, text=text, font=('sans-serif', 11, font.BOLD), fg=textcolor).pack(fill=BOTH, expand=TRUE)
        pop_up_window.mainloop()
#book window
def create_book_window():
    global BOOK_WINDOW_CREATED #make this variable a global variable
    if BOOK_WINDOW_CREATED is not True: #if variable is false
        BOOK_WINDOW_CREATED = True #window created
        #create book window
        book_window = Toplevel()
        #book window closing function
        def close_book_window():
            global BOOK_WINDOW_CREATED #make it a global variable
            BOOK_WINDOW_CREATED = False #set to false so the window can be created again
            book_window.destroy()
        book_window.protocol('WM_DELETE_WINDOW', close_book_window) #amo lang ina sya aram ko na protocol 'WM_DELETE_WINDOW' meaning pag gin close mo sya ica-call nya an 'close_book_window' function
        book_window.title("Book") #title san window
        book_window.resizable(False, False) #disable resizing
        width = 550 #window's width
        height = 500 #window's height
        #screen dimension
        screen_width = book_window.winfo_screenwidth() #screen max width pixels
        screen_height = book_window.winfo_screenheight() #screen max height pixels
        #kailangan ini sya para ma center an aton windows 
        center_x = int(screen_width/2 - width/2)
        center_y = int(screen_height/2 - height/2)
        #tas i-seset naton sya
        book_window.geometry(f'{width}x{height}+{center_x}+{center_y}') #set book window's height and width
        book_window.configure(bg='#FFFFFF') #main window background color
        
        #variables
        duration_price = 430 #per day price
        expected_date = StringVar()
        firstname = StringVar()
        lastname = StringVar()
        address = StringVar()
        contact = StringVar()
        email = StringVar()
        price = StringVar()
        duration = StringVar()
        #set values

        #header text
        Label(book_window, text='Customer Data', font=('sans-serif', 25, font.BOLD), bg='#FFFFFF',fg='#787A40').place(x=170, y=30)
        #first name
        Label(book_window, text='First name:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=0, sticky=E, pady=(115, 5), padx=(100, 0))
        first_name_entry = Entry(book_window, width=15, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=firstname).grid(column=1, row=0, pady=(115, 5), padx=(0, 20))
        #last name
        Label(book_window, text='Last name:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=2, row=0, sticky=E, pady=(115, 5))
        last_name_entry = Entry(book_window, width=15, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=lastname).grid(column=3, row=0, pady=(115, 5))
        #address 
        Label(book_window, text='Address:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=1, sticky=E)
        address_entry = Entry(book_window, width=35, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=address).grid(column=1, row=1, columnspan=3, sticky=W, pady=(5))
        #contact number
        Label(book_window, text='Contact #:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=2, sticky=E)
        contact_number_entry = Entry(book_window, width=35, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=contact).grid(column=1, row=2, columnspan=3, sticky=W, pady=(5))
        #email address
        Label(book_window, text='Email:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=3, sticky=E)
        email_address_entry = Entry(book_window, width=35, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=email).grid(column=1, row=3, columnspan=3, sticky=W, pady=(5))
        #payment and duration
        Label(book_window, text='Expected Date and Duration', font=('sans-serif', 20, font.BOLD), bg='#FFFFFF',fg='#787A40').place(x=110, y=255)
        #format label
        Label(book_window, text='Format: mm/dd/yy', background='#FFFFFF', font=('sans-serif', 9), fg='gray').place(x=170, y=305)
        #expected date
        Label(book_window, text='Expected date:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=4, sticky=E, pady=(90, 5))
        date_entry = Entry(book_window, width=15, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=expected_date)
        date_entry.grid(column=1, row=4, sticky=W, pady=(90, 5), columnspan=3)
        #duration
        Label(book_window, text='Duration:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=5, sticky=E, pady=(5))
        duration_entry = Entry(book_window, width=4, highlightthickness=1, highlightbackground='#c0c4c1', textvariable=duration)
        duration_entry.grid(column=1, row=5, sticky=W, pady=(5))
        #set price function
        def setPrice(event):
            _duration = duration.get() #get entered duration value
            if (_duration.isnumeric()): #if the entered value is numeric
                #set price
                price.set(duration_price * int(_duration)) #price.set(per day price multiply to entered duration)
            if (len(_duration) is 0): 
                price.set("")
        #call setPrice function every time the user presses a key and releases it
        duration_entry.bind('<KeyRelease>', setPrice) #bind a variable (event, function to call)
        Label(book_window, text='day(s)', background='#FFFFFF', font=('sans-serif', 10)).grid(column=1, row=5, sticky=NS, pady=(5))
        #price
        Label(book_window, text='Price:', background='#FFFFFF', font=('sans-serif', 10)).grid(column=0, row=6, sticky=E, pady=(5))
        Label(book_window, textvariable=price, background='#FFFFFF', font=('sans-serif', 10)).grid(column=1, row=6, sticky=W, pady=(5))
        #submit customer's data
        def submit():
            #get variables value
            name = firstname.get().upper() + " " + lastname.get().upper() #store full name as UPPERCASE
            _address = address.get().upper() #store entered address as UPPERCASE
            _contact = contact.get() #get entered contact number
            _email = email.get() #get entered email address
            _expected_date = expected_date.get() #get entered date
            _duration = duration.get() + " day(s)" #
            _amount_paid = price.get() #amount paid to be stored

            #verify if any of the entries has 0 length value
            #list variables to be stored
            variables = [name, _address, _contact, _email, _expected_date, _duration, _amount_paid]
            hasZeroValue = False #boolean variable for verification
            for x in variables: 
                if (len(x) == 0): #if any of the list has 0 length value
                    hasZeroValue = True
            #if none of the list has 0 length value
            if (hasZeroValue is False): 
                global BOOK_WINDOW_CREATED #make this variable a global variable
                #slice or remove the first entered value if it starts with 0
                if (_duration[:1] == '0' or 0):
                    _duration = _duration[1:]
                #append data in the worksheet | excel database
                ws.append([name, _address, _contact, _email, _expected_date, _duration, _amount_paid])
                #get excel max occupied row
                max_row = ws.max_row
                #style occupied cells
                for x in range(1, (max_row + 1)):
                    for y in range(1, 8):
                        ws.cell(row=x, column=y).alignment = Alignment(horizontal='center', vertical='center') #horizontally and vertically center cells
                #save worksheet
                wb.save('customer_database.xlsx')
                #close book window
                BOOK_WINDOW_CREATED = False #set to false so the window can be created again
                book_window.destroy() #close function
                #success submitting data
                title = 'Success' #
                msg = 'Book success.' #
                txtcolor = '#787A40' #
                #create success pop-up window
                create_pop_up_window(title, msg, txtcolor) #function(arguments)
            #if any of the variables list has 0 length value
            else: #error submitting data
                title = 'There seems to be a problem' #
                msg = 'Please fill out the entries.' #
                txtcolor = 'red' #
                #create error pop-up window message
                create_pop_up_window(title, msg, txtcolor) #function(arguments)

        #submit button
        submit_reservation = Button(book_window, text='BOOK', font=('sans-serif', 11, font.BOLD), fg='#C8AB65', bg='#787A40', borderwidth=0, width=35, height=1, command=submit)
        submit_reservation.place(x=120, y=440) #tkinter geometry manager
        book_window.mainloop()

#customers window
def create_customers_window():
    global CUSTOMERS_WINDOW_CREATED #make this variable a global variable
    if CUSTOMERS_WINDOW_CREATED is not True: #if the variable is false
        CUSTOMERS_WINDOW_CREATED = True #set variable to true(window created)
        #create customers window
        customers_window = Toplevel()
        #book window closing function
        def close_customers_window():
            global CUSTOMERS_WINDOW_CREATED #make it a global variable
            CUSTOMERS_WINDOW_CREATED = False #set to false so the window can be created again
            customers_window.destroy() #window close function
        customers_window.protocol('WM_DELETE_WINDOW', close_customers_window) #amo lang ina sya aram ko na protocol 'WM_DELETE_WINDOW' meaning pag gin close mo sya ica-call nya an 'close_customers_window' function
        customers_window.title("Customers") #title san window
        customers_window.resizable(False, False) #disable resizing
        width = 750 #window's width
        height = 450 #window's height
        #screen dimension
        screen_width = customers_window.winfo_screenwidth() #screen max width pixels
        screen_height = customers_window.winfo_screenheight() #screen max height pixels
        #kailangan ini sya para ma center an aton windows 
        center_x = int(screen_width/2 - width/2)
        center_y = int(screen_height/2 - height/2)
        #tas i-seset naton sya
        customers_window.geometry(f'{width}x{height}+{center_x}+{center_y}') #customers window's height and width
        customers_window.configure(bg='#FFFFFF')
        #window's body 
        #make the table scrollable
        lframe = LabelFrame(customers_window)
        canvas = Canvas(lframe, width=750, height=450)
        frame = Frame(canvas)
        scrollbar = Scrollbar(lframe, orient=VERTICAL, command=canvas.yview)
        scrollbar.place(x=730, y=30, height=430)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind('<Configure>', lambda e : canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0,0), window=frame, anchor=NW)
        #get max occupied row in excel
        maxRow = ws.max_row + 1
        #fetch data from excel and store in a list variable
        customersData_list = [] #list variable
        for r in range(2, maxRow): #from range 2, kay dili naton kailangan an row 1 values sa aton excel database(headers lang idto sya)
            customer = {} #create empty dictionary variable
            customer['name'] = ws.cell(row=r, column=1).value #add name key and cell value fetched from excel
            customer['address'] = ws.cell(row=r, column=2).value #add address key and cell value fetched from excel
            customer['contact'] = ws.cell(row=r, column=3).value #add name key and cell value fetched from excel
            customer['email'] = ws.cell(row=r, column=4).value #add name key and cell value fetched from excel
            customer['expected_date'] = ws.cell(row=r, column=5).value #add name key and cell value fetched from excel
            customer['duration'] = ws.cell(row=r, column=6).value #add name key and cell value fetched from excel
            customer['amount_paid'] = ws.cell(row=r, column=7).value #add name key and cell value fetched from excel
            #append created customer list
            customersData_list.append(customer)
        #sort by date
        def sortbydate(e):
            return e['expected_date']
        customersData_list.sort(reverse=True,key=sortbydate)
        _row = 0
        for c in customersData_list: #get listed dictionaries
            #style rows
            bg = '#ffffff' #background color
            fg = '#000' #font(text) color
            if not(_row % 2 == 1): #if _row variable value is an odd number
                bg = '#f0f0ed' #
            customer_data = []
            for v in c: #get every key values in a listed dictionary
                customer_data.append(c[v]) #store value in customer_data list variable
            if len(customer_data[0]) > 15: #if customer's full name length value is greater than 15
                #format name
                name = customer_data[0] 
                _slicedName = name[0:11] + ".." #formatted name
                Label(frame, text=_slicedName, width=15, bg=bg, fg=fg).grid(column=0, row=_row, ipady=2)
            else:
                Label(frame, text=customer_data[0], width=15, bg=bg, fg=fg).grid(column=0, row=_row, ipady=2)
            Label(frame, text=customer_data[1], width=15, bg=bg, fg=fg).grid(column=1, row=_row, ipady=2)
            Label(frame, text=customer_data[2], width=15, bg=bg, fg=fg).grid(column=2, row=_row, ipady=2)
            if len(customer_data[3]) > 15: #if email address length value is greater than 15
                #format email
                email = customer_data[3] 
                _sliced_email = email[0:15] + ".." #formatted email
                Label(frame, text=_sliced_email, width=23, bg=bg, fg=fg).grid(column=3, row=_row, ipady=2)
            else:
                Label(frame, text=customer_data[3], width=23, bg=bg, fg=fg).grid(column=3, row=_row, ipady=2)
            Label(frame, text=customer_data[4], width=10, bg=bg, fg=fg).grid(column=4, row=_row, ipady=2)
            Label(frame, text=customer_data[5], width=10, bg=bg, fg=fg).grid(column=5, row=_row, ipady=2, ipadx=(5))
            Label(frame, text=customer_data[6], width=10, bg=bg, fg=fg).grid(column=6, row=_row, ipady=2)
            _row += 1 #increment row by 1
            
        #headers
        headers = ['Customer Name', 'Address', 'Contact #', 'Email Address', 'Expected Date', 'Duration', 'Amount Paid']
        Label(customers_window, text=headers[0], fg='#C8AB65', bg='#787A40', width=15, height=2).place(x=0, y=0)
        Label(customers_window, text=headers[1], fg='#C8AB65', bg='#787A40', width=17, height=2).place(x=110, y=0)
        Label(customers_window, text=headers[2], fg='#C8AB65', bg='#787A40', width=15, height=2).place(x=230, y=0)
        Label(customers_window, text=headers[3], fg='#C8AB65', bg='#787A40', width=23, height=2).place(x=330, y=0)
        Label(customers_window, text=headers[4], fg='#C8AB65', bg='#787A40', width=12, height=2).place(x=496, y=0)
        Label(customers_window, text=headers[5], fg='#C8AB65', bg='#787A40', width=12, height=2).place(x=580, y=0)
        Label(customers_window, text=headers[6], fg='#C8AB65', bg='#787A40', width=13, height=2).place(x=660, y=0)
        canvas.grid(column=0, row=0, pady=(40, 5))
        lframe.pack()
        customers_window.mainloop()

#create main window
main = Tk() 
main.title("Reservation System") #title san window
main.resizable(False, False) #disable resizing
width = 600 #window's width
height = 450 #window's height
#screen dimension
screen_width = main.winfo_screenwidth() #screen max width pixels
screen_height = main.winfo_screenheight() #screen max height pixels
#kailangan ini sya para ma center an aton windows
center_x = int(screen_width/2 - width/2)
center_y = int(screen_height/2 - height/2)
#tas i-seset naton sya
main.geometry(f'{width}x{height}+{center_x}+{center_y}') #set main window's height and width
main.configure(bg='#FFFFFF') #main window background color
#display time
def start_timer(): #timer loop function
    current_time = strftime("%I:%M:%S") #current time, an strftime(), format ina sya san %HOUR %MINUTES %SECONDS https://www.w3schools.com/python/python_datetime.asp
    time.configure(text=current_time) #tas configure() meaning i-seset or i-change an text value san aton time label widget (inan sa baba)
    time.after(1000, start_timer) #after method an first argument milliseconds ina 1000 ms = 1 second, tas an panduwa an start_timer function ica-call sya para mag loop after san 1000 milliseconds
#time label widget
time = Label(main, font=('sans-serif', 15, font.BOLD), fg='#787A40', bg='#FFFFFF') #(container, font, text color, background-color)
time.grid(column=0, row=0, ipadx=20, ipady=20) #grid geometry manager, 3 ina sya (grid, pack, place) https://www.pythontutorial.net/tkinter/tkinter-pack/
#book button
book_button = Button(main, text='BOOK', font=('sans-serif', 11, font.BOLD), fg='#C8AB65', bg='#787A40', borderwidth=0, width=10, height=2, command=create_book_window) #(container, text value, font(font-family, size, weight), text color, background color, call create_book_window function)
book_button.place(x=190, y=220) #place geometry manager, (x, y) specific coordinates
#customers button
customers_button = Button(main, text='CUSTOMERS', font=('sans-serif', 11, font.BOLD), fg='#C8AB65', bg='#787A40', borderwidth=0, width=15, height=2, command=create_customers_window) #(container, text value, font(font-family, size, weight), text color, background color, call create_customers_window function)
customers_button.place(x=300, y=220) #place geometry manager, (x, y) specific coordinates
#call start timer
start_timer()
main.mainloop() #mainloop() need ini sya para ma display an window